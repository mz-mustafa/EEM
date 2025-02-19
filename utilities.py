from zipfile import BadZipFile
from scenario_sorted import Scenario
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd
#from requests_oauthlib import OAuth2Session
import os
#import io
from sources import GridSource, SolarSource, WindSource, GasGenSource, \
    HFOGenSource, TrifuelGenSource, BESSSource, DieselGenSource, PPASource, ExistingGasGenSource

def has_duplicate_values(var_list):
    filtered_list = [item for item in var_list if item != -1]

    seen = set()
    for value in filtered_list:
        if value in seen:
            return True
        seen.add(value)
    return False


def get_dataframes_and_sheets(sc):
    dataframes = [sc.scenario_spec,
                  sc.power_df,
                  sc.energy_df,
                  sc.capex_df,
                  sc.opex_df,
                  sc.emissions_df,
                  sc.summary_df,
                  sc.power_summary_df,
                  sc.energy_summary_df,
                  sc.energy_summary_concise_df,
                  sc.emissions_summary_df,
                  sc.opex_summary_df,
                  sc.opex_summary_concise_df
                  ]
    sheets = ['scenario', 'power', 'energy', 'capex', 'opex', 'emissions',
              'summary', 'power_summary', 'energy_summary', 'energy_summary_concise',
              'emissions_summary', 'opex_summary', 'opex_summary_concise']
    return dataframes, sheets


def write_df_to_excel_sheet(df, sheet_name, file_name='outputs.xlsx'):
    # Load the workbook
    try:
        book = load_workbook(file_name)
    except BadZipFile as e:
        print(f"Error opening file: {e}")
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    book.save(file_name)

    # Now, write the DataFrame to the Excel sheet using pandas
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def write_results_to_outputs(sc):
    dataframes, sheets = get_dataframes_and_sheets(sc)
    for df, sheet in zip(dataframes, sheets):
        write_df_to_excel_sheet(df, sheet)

def generate_excel_in_memory(sc):
    dataframes, sheets = get_dataframes_and_sheets(sc)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for df, sheet in zip(dataframes, sheets):
            df.to_excel(writer, sheet_name=sheet, index=False)
    output.seek(0)
    return output

##TEST CODE
#create scenario, extract inputs
#session = connect_onedrive()
#get_onedrive_file_list(session)
#read_file(session, file_format=True)
n = 8
sc = Scenario('','Pakistan Cables Limited','input_data.xlsx',n)


#GRID CONFIG
grid_source = GridSource(n,3)
grid_source.inputs[0]['count_prim_units'] = 1
grid_source.inputs[0]['rating_prim_units'] = 4.5

#PPA CONFIG
ppa_source = PPASource(n,0)
ppa_source.inputs[1]['count_prim_units'] = 1
ppa_source.inputs[1]['rating_prim_units'] = 2


#SOLAR CONFIG
solar_source = SolarSource(n,0)

# Update the input structure for Year 0
solar_source.inputs[0]['count_prim_units'] = 1
solar_source.inputs[0]['rating_prim_units'] = 2

#GAS CONFIG
gas_gen_source = GasGenSource(n,1)

# Update the input structure for Year 0
gas_gen_source.inputs[0]['count_prim_units'] = 1
gas_gen_source.inputs[0]['rating_prim_units'] = 1.5
gas_gen_source.inputs[0]['perc_rated_output'] = 100
gas_gen_source.inputs[0]['fuel_eff'] = 100

# Update chp_operation and gas_fuel_type values
gas_gen_source.inputs['chp_operation'] = True
gas_gen_source.inputs['fuel_type'] = 'RLNG'


#EXISTING GAS CONFIG
existing_gas_gen_source = ExistingGasGenSource(n,1)

existing_gas_gen_source.inputs[2]['count_prim_units'] = 2
existing_gas_gen_source.inputs[2]['rating_prim_units'] = 1
existing_gas_gen_source.inputs[2]['perc_rated_output'] = 90
existing_gas_gen_source.inputs[2]['fuel_eff'] = 90

existing_gas_gen_source.inputs['chp_operation'] = True
existing_gas_gen_source.inputs['fuel_type'] = 'NG'


#BESS CONFIG
bess_source = BESSSource(n,0)
bess_source.inputs[1]['count_prim_units'] = 3
bess_source.inputs[1]['rating_prim_units'] = 0.5


#WIND CONFIG
wind_source = WindSource(n,0)
wind_source.inputs[2]['count_prim_units'] = 1
wind_source.inputs[2]['rating_prim_units'] = 2


#HFO CONFIG
hfo_gen_source = HFOGenSource(n,1)

# Update the input structure for Year 2
hfo_gen_source.inputs[1]['count_prim_units'] = 1
hfo_gen_source.inputs[1]['rating_prim_units'] = 1.5
hfo_gen_source.inputs[1]['perc_rated_output'] = 100
hfo_gen_source.inputs[1]['fuel_eff'] = 100

#Trifuel CONFIG
tf_gen_source = TrifuelGenSource(n,1)

# Update the input structure for Year 2
tf_gen_source.inputs[1]['count_prim_units'] = 1
tf_gen_source.inputs[1]['rating_prim_units'] = 1.5
tf_gen_source.inputs[1]['perc_rated_output'] = 100
tf_gen_source.inputs[1]['fuel_eff'] = 100

#DIESEL CONFIG
dg_source = DieselGenSource(n,4)

# Update the input structure for Year 1
dg_source.inputs[1]['count_prim_units'] = 1
dg_source.inputs[1]['rating_prim_units'] = 1.2
dg_source.inputs[1]['perc_rated_output'] = 100
dg_source.inputs[1]['fuel_eff'] = 100

dg_source.inputs[4]['count_prim_units'] = 1
dg_source.inputs[4]['rating_prim_units'] = 1.2
dg_source.inputs[4]['perc_rated_output'] = 100
dg_source.inputs[4]['fuel_eff'] = 100


sc.add_source(grid_source)
print(f"{sc.sources_dict['Grid'].source_type} added.")

#sc.add_source(ppa_source)
#print(f"{sc.sources_dict['PPA'].source_type} added.")

sc.add_source(solar_source)
print(f"{sc.sources_dict['Solar'].source_type} added.")

sc.add_source(gas_gen_source)
print(f"{sc.sources_dict['Gas Generator'].source_type} added.")

sc.add_source(existing_gas_gen_source)
print(f"{sc.sources_dict['Existing Gas Generators'].source_type} added.")

#sc.add_source(wind_source)
#print(f"{sc.sources_dict['Wind'].source_type} added.")

#sc.add_source(hfo_gen_source)
#print(f"{sc.sources_dict['HFO Generator'].source_type} added.")

#sc.add_source(tf_gen_source)
#print(f"{sc.sources_dict['HFO+Gas Generator'].source_type} added.")


sc.add_source(bess_source)
print(f"{sc.sources_dict['BESS'].source_type} added.")

sc.add_source(dg_source)
print(f"{sc.sources_dict['Diesel Generator'].source_type} added.")


sc.generate_results()
sc.generate_summaries()
write_results_to_outputs(sc)

## TEST CODE END